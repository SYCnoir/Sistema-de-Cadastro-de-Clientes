#importação das bibliotecas necessarias
from InquirerPy import inquirer 
from rich.console import Console 
from rich.table import Table
from rich.panel import Panel
from rich.rule import Rule
import openpyxl
import os 

#console do rich para exibir os elementos visuais no terminal
console = Console()

#nome do arquivo Excel onde os dados serão salvos
ARQUIVO_EXCEL = 'Clientes.xlsx'

def carregar_clientes():
    #carrega os clientes do arquivo excel ao inciar o sistema
    clientes = []
    if not os.path.exists(ARQUIVO_EXCEL):
        return clientes
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
    ws = wb.active 
    for linha in ws.iter_rows(min_row=2, values_only=True):
        clientes.append({
            'nome': linha[0],
            'cpf': linha[1],
            'email': linha[2],
            'telefone': linha[3]
        })
    return clientes 

def salvar_clientes(clientes):
    #salva todos os clientes no arquivo excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    #cabeçalho
    ws.append(['Nome', 'CPF', 'Email', 'Telefone'])
    #dados
    for cliente in clientes:
        ws.append([cliente['nome'], cliente['cpf'], cliente['email'], cliente['telefone']])
    wb.save(ARQUIVO_EXCEL)

def cadastrar_cliente(clientes):
    #coleta os dados do novo cliente com validação
    console.print(Rule('Cadastrar Cliente', style='green'))
    nome = inquirer.text(
        message= 'Nome completo:',
        validate=lambda x: len(x) > 2 or 'Nome muito curto!'
    ).execute()
    cpf = inquirer.text(
        message='CPF:',
        validate=lambda x: len(x) == 11 or 'CPF deve ter 11 dígitos!'
    ).execute()
    email = inquirer.text(
        message='Email:', 
        validate=lambda x: '@' in x or 'Email inválido!'
    ).execute()
    telefone = inquirer.text(
        message='Telefone:',
        validate=lambda x: len(x) >= 10 or 'Telefone inválido!'
    ).execute()
    clientes.append({
    'nome': nome,
    'cpf': cpf,
    'email': email,
    'telefone': telefone,
     })  
    salvar_clientes(clientes)
    console.print(Panel(f'Cliente [bold]{nome}[/bold] cadastrado com sucesso!', style='green'))

def listar_clientes(clientes):
    #exibe todos os clientes em uma tabela formatada com rich
    console.print(Rule('Lista de Clientes', style='blue'))
    if not clientes:
        console.print(Panel('Nenhum cliente cadastrado!', style='yellow'))
        return
    tabela = Table(title=' Clientes cadastrados', show_header=True, header_style='bold blue')
    tabela.add_column('Nome', style='cyan')
    tabela.add_column('CPF', style='white')
    tabela.add_column('Email', style='white')
    tabela.add_column('Telefone', style='white')
    for cliente in clientes:
        tabela.add_row(cliente['nome'], cliente['cpf'], cliente['email'], cliente['telefone'])
    console.print(tabela)

def buscar_cliente(clientes):
    #busca um cliente pelo nome
    console.print(Rule('Busca Cliente', style='yellow'))
    nome = inquirer.text(message='Digite o nome para buscar:').execute()
    encontrados = [c for c in clientes if nome.lower() in c ['nome'].lower()]
    if not encontrados:
        console.print(Panel('Nenhum cliente encontrado!', style='red'))
        return
    tabela = Table(title='Resultado da Busca', show_header=True, header_style='bold yellow')
    tabela.add_column('Nome', style='cyan')
    tabela.add_column('CPF', style='white')
    tabela.add_column('Email', style='white')
    tabela.add_column('Telefone', style='white')
    for cliente in encontrados:
        tabela.add_row(cliente['nome'], cliente['cpf'], cliente['email'], cliente['telefone'])
    console.print(tabela)

def excluir_cliente(clientes):
    # remove um cliente da lista pelo nome
    console.print(Rule('Excluir Cliente', style='red'))
    nome = inquirer.text(message='Digite o nome do cliente a excluir:').execute()
    encontrados = [c for c in clientes if nome.lower() in c['nome'].lower()]
    if not encontrados:
        console.print(Panel('Cliente não encontrado!', style='red'))
        return
    clientes.remove(encontrados[0])
    salvar_clientes(clientes)
    console.print(Panel(f'Cliente [bold]{encontrados[0]["nome"]}[/bold] excluído!', style='red'))
      
def menu():
    #menu principal do sistema com loop while
    clientes = carregar_clientes()
    while True:
        console.print(Panel('[bold cyan]Sistema de Cadastro de Clientes[/bold cyan]', style='cyan'))
        opcao = inquirer.select(
            message='Escolha uma opção:',
            choices=[
                '1 - Cadastrar cliente',
                '2 - Listar clientes',
                '3 - Buscar cliente',
                '4 - Excluir cliente',
                '5 - Sair'
            ]
        ).execute()
        if opcao == '1 - Cadastrar cliente':
            cadastrar_cliente(clientes)
        elif opcao == '2 - Listar clientes':
            listar_clientes(clientes)
        elif opcao == '3 - Buscar cliente':
             buscar_cliente(clientes)
        elif opcao == '4 - Excluir cliente':
            excluir_cliente(clientes)
        elif opcao == '5 - Sair':
            console.print(Panel('Saindo do sistema!', style='red'))
            break
    

menu()



